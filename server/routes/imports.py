from fastapi import APIRouter, HTTPException, UploadFile, File
from server.db import pool
from server.models import ImportData, ImportValidate
from server.routes.data import _cast_value
from server.validators import validate_with_code
import io
import openpyxl

router = APIRouter(tags=["imports"])


@router.post("/cruds/{crud_id}/import/preview")
async def preview_excel(crud_id: int, file: UploadFile = File(...)):
    """Read Excel file and return column names + sample data for mapping UI."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Apenas arquivos .xlsx sao suportados")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(400, "Planilha vazia")

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        raise HTTPException(400, "Nenhum cabecalho encontrado")

    headers = [str(h).strip() if h else f"Coluna_{i}" for i, h in enumerate(headers)]

    sample_data = []
    for i, row in enumerate(rows_iter):
        if i >= 5:
            break
        sample_data.append([str(v) if v is not None else "" for v in row])

    total_rows = ws.max_row - 1 if ws.max_row else 0
    wb.close()

    return {
        "file_name": file.filename,
        "headers": headers,
        "sample_data": sample_data,
        "total_rows": total_rows,
    }


def _get_columns_meta(cur, crud_id: int):
    """Get full column metadata for validation."""
    cur.execute(
        """SELECT db_column, name, data_type, is_required,
                  COALESCE(is_unique, FALSE), COALESCE(validation_rule, '')
           FROM crud_columns WHERE crud_id = %s AND is_deleted = FALSE ORDER BY position""",
        (crud_id,),
    )
    return [
        {"db_column": r[0], "name": r[1], "data_type": r[2],
         "is_required": r[3], "is_unique": r[4], "validation_rule": r[5]}
        for r in cur.fetchall()
    ]


@router.post("/cruds/{crud_id}/import/validate")
def validate_import(crud_id: int, body: ImportValidate):
    """Validate all rows before importing. Checks: required, unique, AI rules."""
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT table_name FROM crud_definitions WHERE id = %s AND is_deleted = FALSE",
                (crud_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Tabela nao encontrada")
            table_name = row[0]

            columns = _get_columns_meta(cur, crud_id)
            col_map = {c["db_column"]: c for c in columns}

            # Get existing values for unique columns
            existing_unique = {}
            for col in columns:
                if col["is_unique"] and col["db_column"] in body.mapping:
                    cur.execute(
                        f'SELECT "{col["db_column"]}" FROM "{table_name}" WHERE _is_deleted = FALSE'
                    )
                    existing_unique[col["db_column"]] = {str(r[0]) for r in cur.fetchall() if r[0] is not None}

    # Validate each row
    results = []
    seen_unique = {col["db_column"]: set() for col in columns if col["is_unique"]}

    for i, excel_row in enumerate(body.rows):
        errors = []

        for crud_col, excel_col in body.mapping.items():
            if crud_col not in col_map:
                continue
            col = col_map[crud_col]
            raw_val = excel_row.get(excel_col)
            val = str(raw_val).strip() if raw_val is not None else ""

            # Required check
            if col["is_required"] and not val:
                errors.append({"field": col["name"], "message": "Campo obrigatorio"})
                continue

            if not val:
                continue

            # Unique check (against DB + within import batch)
            if col["is_unique"]:
                if val in existing_unique.get(crud_col, set()):
                    errors.append({"field": col["name"], "message": f"Valor ja existe no banco: {val}"})
                elif val in seen_unique.get(crud_col, set()):
                    errors.append({"field": col["name"], "message": f"Valor duplicado na planilha: {val}"})
                else:
                    seen_unique.setdefault(crud_col, set()).add(val)

            # AI validation rule
            if col["validation_rule"] and val:
                try:
                    ai_result = validate_with_code(val, col["validation_rule"], col["name"], model=body.model)
                    if not ai_result.get("valid"):
                        errors.append({"field": col["name"], "message": ai_result.get("message", "Invalido")})
                except Exception:
                    pass

        results.append({
            "row": i + 1,
            "valid": len(errors) == 0,
            "errors": errors,
            "preview": {col_map[k]["name"]: str(excel_row.get(v, ""))[:50]
                        for k, v in body.mapping.items() if k in col_map},
        })

    valid_count = sum(1 for r in results if r["valid"])
    return {
        "total": len(results),
        "valid": valid_count,
        "invalid": len(results) - valid_count,
        "rows": results,
    }


@router.post("/cruds/{crud_id}/import")
def import_data(crud_id: int, body: ImportData):
    """Import mapped data into the CRUD table."""
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT table_name FROM crud_definitions WHERE id = %s AND is_deleted = FALSE",
                (crud_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Tabela nao encontrada")
            table_name = row[0]

            cur.execute(
                """SELECT db_column, data_type FROM crud_columns
                   WHERE crud_id = %s AND is_deleted = FALSE""",
                (crud_id,),
            )
            col_types = {r[0]: r[1] for r in cur.fetchall()}

            valid_mappings = {}
            for crud_col, excel_col in body.mapping.items():
                if crud_col in col_types and excel_col:
                    valid_mappings[crud_col] = excel_col

            if not valid_mappings:
                raise HTTPException(400, "Nenhum mapeamento valido")

            inserted = 0
            errors = 0

            for excel_row in body.rows:
                insert_cols = []
                insert_vals = []
                params = []

                for crud_col, excel_col in valid_mappings.items():
                    val = excel_row.get(excel_col)
                    insert_cols.append(f'"{crud_col}"')
                    insert_vals.append("%s")
                    params.append(_cast_value(val, col_types[crud_col]))

                try:
                    cur.execute(
                        f"""INSERT INTO "{table_name}" ({', '.join(insert_cols)})
                            VALUES ({', '.join(insert_vals)})""",
                        params,
                    )
                    inserted += 1
                except Exception:
                    errors += 1

            conn.commit()

    return {
        "inserted": inserted,
        "errors": errors,
        "message": f"{inserted} registros importados com sucesso"
        + (f", {errors} com erro" if errors else ""),
    }


@router.get("/cruds/{crud_id}/export")
def export_data(crud_id: int):
    """Export all data as JSON (frontend converts to Excel)."""
    with pool.connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT table_name FROM crud_definitions WHERE id = %s AND is_deleted = FALSE",
                (crud_id,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "Tabela nao encontrada")
            table_name = row[0]

            cur.execute(
                """SELECT name, db_column, data_type FROM crud_columns
                   WHERE crud_id = %s AND is_deleted = FALSE ORDER BY position""",
                (crud_id,),
            )
            columns = [{"name": r[0], "db_column": r[1], "data_type": r[2]} for r in cur.fetchall()]

            db_cols = [c["db_column"] for c in columns]
            select_cols = ", ".join(f'"{c}"' for c in db_cols)

            cur.execute(
                f'SELECT _id, {select_cols} FROM "{table_name}" WHERE _is_deleted = FALSE ORDER BY _id'
            )
            rows = cur.fetchall()

    data = []
    for row in rows:
        item = {}
        for i, col in enumerate(columns):
            val = row[1 + i]
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            item[col["name"]] = val
        data.append(item)

    return {"columns": [c["name"] for c in columns], "data": data}
